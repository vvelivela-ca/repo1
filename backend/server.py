from fastapi import FastAPI, APIRouter, HTTPException, UploadFile, File, Form
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone
import yfinance as yf
import csv
import io
import re
import pdfplumber
import json as json_module
from openpyxl import load_workbook
from emergentintegrations.llm.chat import LlmChat, UserMessage

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ── Asset Types & Exchange Detection ────────────────────
ASSET_TYPES = ["Stock", "ETF", "Mutual Fund", "Crypto", "Bond", "Real Estate", "Mortgage", "Cash", "Other"]

# Known exchange suffixes for yfinance
EXCHANGE_SUFFIXES = {
    # North America
    "NYSE": "",
    "NASDAQ": "",
    "AMEX": "",
    "TSX": ".TO",      # Toronto Stock Exchange (Canada)
    "TSXV": ".V",      # TSX Venture Exchange
    "CSE": ".CN",      # Canadian Securities Exchange
    # India
    "NSE": ".NS",      # National Stock Exchange of India
    "BSE": ".BO",      # Bombay Stock Exchange
    # Europe
    "LSE": ".L",       # London Stock Exchange
    "XETRA": ".DE",    # German Exchange
    "EURONEXT": ".PA", # Paris
    # Asia
    "HKEX": ".HK",     # Hong Kong
    "SGX": ".SI",      # Singapore
    "ASX": ".AX",      # Australia
    # Crypto
    "CRYPTO": "-USD",
}

# Currency by exchange
EXCHANGE_CURRENCIES = {
    "NYSE": "USD", "NASDAQ": "USD", "AMEX": "USD",
    "TSX": "CAD", "TSXV": "CAD", "CSE": "CAD",
    "NSE": "INR", "BSE": "INR",
    "LSE": "GBP", "XETRA": "EUR", "EURONEXT": "EUR",
    "HKEX": "HKD", "SGX": "SGD", "ASX": "AUD",
    "CRYPTO": "USD",
}

# Common crypto symbols
CRYPTO_SYMBOLS = {"BTC", "ETH", "XRP", "SOL", "DOGE", "ADA", "DOT", "MATIC", "LINK", "AVAX", "SHIB", "LTC", "UNI", "ATOM", "XLM"}

# ── Models ──────────────────────────────────────────────

class PortfolioCreate(BaseModel):
    name: str
    portfolio_type: str = "Investment"

class PortfolioUpdate(BaseModel):
    name: Optional[str] = None
    portfolio_type: Optional[str] = None

class PortfolioResponse(BaseModel):
    id: str
    name: str
    portfolio_type: str
    created_at: str

class HoldingCreate(BaseModel):
    symbol: str
    shares: float
    avg_price: float
    portfolio_id: str
    currency: Optional[str] = None  # Now optional - will auto-detect
    asset_type: Optional[str] = None  # Now optional - will auto-detect
    exchange: Optional[str] = None
    notes: Optional[str] = None

class HoldingUpdate(BaseModel):
    symbol: Optional[str] = None
    shares: Optional[float] = None
    avg_price: Optional[float] = None
    currency: Optional[str] = None
    asset_type: Optional[str] = None
    exchange: Optional[str] = None
    notes: Optional[str] = None

class HoldingResponse(BaseModel):
    id: str
    symbol: str
    shares: float
    avg_price: float
    portfolio_id: str
    currency: str
    asset_type: str
    exchange: Optional[str]
    notes: Optional[str]
    created_at: str
    updated_at: str

class MortgageCreate(BaseModel):
    name: str
    portfolio_id: str
    principal: float
    interest_rate: float
    term_years: int
    start_date: str
    currency: str = "USD"
    property_value: Optional[float] = None
    notes: Optional[str] = None

class MortgageUpdate(BaseModel):
    name: Optional[str] = None
    principal: Optional[float] = None
    interest_rate: Optional[float] = None
    term_years: Optional[int] = None
    start_date: Optional[str] = None
    currency: Optional[str] = None
    property_value: Optional[float] = None
    extra_payments: Optional[float] = None
    notes: Optional[str] = None

class MortgageResponse(BaseModel):
    id: str
    name: str
    portfolio_id: str
    principal: float
    interest_rate: float
    term_years: int
    start_date: str
    currency: str
    property_value: Optional[float]
    extra_payments: float
    notes: Optional[str]
    created_at: str
    updated_at: str

# ── Smart Ticker Detection ──────────────────────────────

async def detect_ticker_info(symbol: str) -> Dict[str, Any]:
    """
    Detect ticker info (exchange, currency, asset type) by trying different exchanges.
    Returns best match with price data.
    """
    symbol = symbol.upper().strip()
    
    # If symbol already has exchange suffix, respect it
    suffix_map = {
        ".TO": ("TSX", "CAD"),
        ".V": ("TSXV", "CAD"),
        ".CN": ("CSE", "CAD"),
        ".NS": ("NSE", "INR"),
        ".BO": ("BSE", "INR"),
        ".L": ("LSE", "GBP"),
        ".DE": ("XETRA", "EUR"),
        ".PA": ("EURONEXT", "EUR"),
        ".HK": ("HKEX", "HKD"),
        ".SI": ("SGX", "SGD"),
        ".AX": ("ASX", "AUD"),
    }
    
    for suffix, (exchange, currency) in suffix_map.items():
        if symbol.endswith(suffix):
            try:
                ticker = yf.Ticker(symbol)
                info = ticker.fast_info
                if hasattr(info, 'last_price') and info.last_price and info.last_price > 0:
                    # Detect asset type
                    asset_type = "Stock"
                    try:
                        ticker_info = ticker.info
                        quote_type = ticker_info.get("quoteType", "").upper()
                        if quote_type == "ETF":
                            asset_type = "ETF"
                        elif quote_type == "MUTUALFUND":
                            asset_type = "Mutual Fund"
                    except:
                        pass
                    return {
                        "symbol": symbol,
                        "exchange": exchange,
                        "currency": currency,
                        "asset_type": asset_type,
                        "price": float(info.last_price)
                    }
            except:
                pass
    
    # Check if it's crypto
    base_symbol = symbol.replace("-USD", "").replace("-CAD", "").replace("-INR", "")
    if base_symbol in CRYPTO_SYMBOLS or "-USD" in symbol or "-CAD" in symbol:
        crypto_symbol = f"{base_symbol}-USD" if "-" not in symbol else symbol
        try:
            ticker = yf.Ticker(crypto_symbol)
            info = ticker.fast_info
            if hasattr(info, 'last_price') and info.last_price and info.last_price > 0:
                return {
                    "symbol": crypto_symbol,
                    "exchange": "CRYPTO",
                    "currency": "USD",
                    "asset_type": "Crypto",
                    "price": float(info.last_price)
                }
        except:
            pass
    
    # Try US markets first (no suffix)
    try:
        ticker = yf.Ticker(symbol)
        info = ticker.fast_info
        if hasattr(info, 'last_price') and info.last_price and info.last_price > 0:
            currency = "USD"
            asset_type = "Stock"
            try:
                ticker_info = ticker.info
                currency = ticker_info.get("currency", "USD").upper()
                quote_type = ticker_info.get("quoteType", "").upper()
                if quote_type == "ETF":
                    asset_type = "ETF"
                elif quote_type == "MUTUALFUND":
                    asset_type = "Mutual Fund"
            except:
                pass
            return {
                "symbol": symbol,
                "exchange": "NYSE/NASDAQ",
                "currency": currency,
                "asset_type": asset_type,
                "price": float(info.last_price)
            }
    except:
        pass
    
    # Try other exchanges if US didn't work
    exchanges_to_try = [
        (".TO", "TSX", "CAD"),
        (".NS", "NSE", "INR"),
        (".L", "LSE", "GBP"),
        (".DE", "XETRA", "EUR"),
    ]
    
    for suffix, exchange, default_currency in exchanges_to_try:
        try_symbol = f"{symbol}{suffix}"
        try:
            ticker = yf.Ticker(try_symbol)
            info = ticker.fast_info
            if hasattr(info, 'last_price') and info.last_price and info.last_price > 0:
                currency = default_currency
                asset_type = "Stock"
                try:
                    ticker_info = ticker.info
                    currency = ticker_info.get("currency", default_currency).upper()
                    quote_type = ticker_info.get("quoteType", "").upper()
                    if quote_type == "ETF":
                        asset_type = "ETF"
                    elif quote_type == "MUTUALFUND":
                        asset_type = "Mutual Fund"
                except:
                    pass
                return {
                    "symbol": try_symbol,
                    "exchange": exchange,
                    "currency": currency,
                    "asset_type": asset_type,
                    "price": float(info.last_price)
                }
        except:
            continue
    
    # Default fallback
    return {
        "symbol": symbol,
        "exchange": None,
        "currency": "USD",
        "asset_type": "Stock",
        "price": 0
    }

# ── Seed Data ───────────────────────────────────────────

SEED_HOLDINGS = [
    {"symbol": "AAPL", "shares": 65, "avg_price": 105.88},
    {"symbol": "QQQ", "shares": 75, "avg_price": 371.92},
    {"symbol": "TSLA", "shares": 100, "avg_price": 161.64},
    {"symbol": "MSFT", "shares": 25, "avg_price": 329.72},
    {"symbol": "GOOGL", "shares": 60, "avg_price": 130.33},
    {"symbol": "BTC-USD", "shares": 0.5, "avg_price": 42000},
    {"symbol": "ETH-USD", "shares": 5, "avg_price": 2200},
]

@app.on_event("startup")
async def seed_data():
    # Migrate existing holdings
    await db.holdings.update_many({"currency": {"$exists": False}}, {"$set": {"currency": "USD"}})
    await db.holdings.update_many({"asset_type": {"$exists": False}}, {"$set": {"asset_type": "Stock"}})
    await db.holdings.update_many({"exchange": {"$exists": False}}, {"$set": {"exchange": None}})
    await db.holdings.update_many({"notes": {"$exists": False}}, {"$set": {"notes": None}})
    await db.portfolios.update_many({"portfolio_type": {"$exists": False}}, {"$set": {"portfolio_type": "Investment"}})

    portfolio_count = await db.portfolios.count_documents({})
    if portfolio_count == 0:
        logger.info("Seeding initial data...")
        now = datetime.now(timezone.utc).isoformat()
        default_id = str(uuid.uuid4())
        await db.portfolios.insert_one({
            "id": default_id,
            "name": "My Portfolio",
            "portfolio_type": "Investment",
            "created_at": now,
        })
        for h in SEED_HOLDINGS:
            # Auto-detect ticker info
            ticker_info = await detect_ticker_info(h["symbol"])
            doc = {
                "id": str(uuid.uuid4()),
                "symbol": ticker_info["symbol"],
                "shares": h["shares"],
                "avg_price": h["avg_price"],
                "currency": ticker_info["currency"],
                "asset_type": ticker_info["asset_type"],
                "exchange": ticker_info["exchange"],
                "notes": None,
                "portfolio_id": default_id,
                "created_at": now,
                "updated_at": now,
            }
            await db.holdings.insert_one(doc)
        logger.info(f"Seeded 1 portfolio + {len(SEED_HOLDINGS)} holdings")

# ── Portfolio CRUD ──────────────────────────────────────

@api_router.get("/portfolios", response_model=List[PortfolioResponse])
async def get_portfolios():
    docs = await db.portfolios.find({}, {"_id": 0}).to_list(50)
    return docs

@api_router.post("/portfolios", response_model=PortfolioResponse)
async def create_portfolio(data: PortfolioCreate):
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.name.strip(),
        "portfolio_type": data.portfolio_type,
        "created_at": now,
    }
    await db.portfolios.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/portfolios/{portfolio_id}", response_model=PortfolioResponse)
async def update_portfolio(portfolio_id: str, data: PortfolioUpdate):
    existing = await db.portfolios.find_one({"id": portfolio_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    update_fields = {}
    if data.name is not None:
        update_fields["name"] = data.name.strip()
    if data.portfolio_type is not None:
        update_fields["portfolio_type"] = data.portfolio_type
    if update_fields:
        await db.portfolios.update_one({"id": portfolio_id}, {"$set": update_fields})
    updated = await db.portfolios.find_one({"id": portfolio_id}, {"_id": 0})
    return updated

@api_router.delete("/portfolios/{portfolio_id}")
async def delete_portfolio(portfolio_id: str):
    count = await db.portfolios.count_documents({})
    if count <= 1:
        raise HTTPException(status_code=400, detail="Cannot delete the last portfolio")
    result = await db.portfolios.delete_one({"id": portfolio_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    await db.holdings.delete_many({"portfolio_id": portfolio_id})
    await db.mortgages.delete_many({"portfolio_id": portfolio_id})
    return {"message": "Portfolio and its holdings deleted"}

# ── Holdings CRUD with Auto-Detection ───────────────────

@api_router.get("/holdings", response_model=List[HoldingResponse])
async def get_holdings(portfolio_id: Optional[str] = None, asset_type: Optional[str] = None):
    query = {}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    if asset_type:
        query["asset_type"] = asset_type
    docs = await db.holdings.find(query, {"_id": 0}).to_list(500)
    return docs

@api_router.post("/holdings", response_model=HoldingResponse)
async def create_holding(data: HoldingCreate):
    portfolio = await db.portfolios.find_one({"id": data.portfolio_id}, {"_id": 0})
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    
    # Auto-detect ticker info if not provided
    ticker_info = await detect_ticker_info(data.symbol)
    
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "symbol": ticker_info["symbol"],
        "shares": data.shares,
        "avg_price": data.avg_price,
        "currency": data.currency.upper().strip() if data.currency else ticker_info["currency"],
        "asset_type": data.asset_type if data.asset_type else ticker_info["asset_type"],
        "exchange": data.exchange if data.exchange else ticker_info["exchange"],
        "notes": data.notes,
        "portfolio_id": data.portfolio_id,
        "created_at": now,
        "updated_at": now,
    }
    await db.holdings.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/holdings/{holding_id}", response_model=HoldingResponse)
async def update_holding(holding_id: str, data: HoldingUpdate):
    existing = await db.holdings.find_one({"id": holding_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Holding not found")
    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    if data.symbol is not None:
        update_fields["symbol"] = data.symbol.upper().strip()
    if data.shares is not None:
        update_fields["shares"] = data.shares
    if data.avg_price is not None:
        update_fields["avg_price"] = data.avg_price
    if data.currency is not None:
        update_fields["currency"] = data.currency.upper().strip()
    if data.asset_type is not None:
        update_fields["asset_type"] = data.asset_type
    if data.exchange is not None:
        update_fields["exchange"] = data.exchange
    if data.notes is not None:
        update_fields["notes"] = data.notes
    await db.holdings.update_one({"id": holding_id}, {"$set": update_fields})
    updated = await db.holdings.find_one({"id": holding_id}, {"_id": 0})
    return updated

@api_router.delete("/holdings/{holding_id}")
async def delete_holding(holding_id: str):
    result = await db.holdings.delete_one({"id": holding_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Holding not found")
    return {"message": "Holding deleted"}

# ── Ticker Lookup API ───────────────────────────────────

@api_router.get("/ticker/lookup/{symbol}")
async def lookup_ticker(symbol: str):
    """Look up a ticker and auto-detect exchange, currency, and asset type."""
    info = await detect_ticker_info(symbol)
    return info

# ── Mortgage CRUD ───────────────────────────────────────

@api_router.get("/mortgages")
async def get_mortgages(portfolio_id: Optional[str] = None):
    query = {}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    docs = await db.mortgages.find(query, {"_id": 0}).to_list(100)
    return docs

@api_router.post("/mortgages", response_model=MortgageResponse)
async def create_mortgage(data: MortgageCreate):
    portfolio = await db.portfolios.find_one({"id": data.portfolio_id}, {"_id": 0})
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")
    now = datetime.now(timezone.utc).isoformat()
    doc = {
        "id": str(uuid.uuid4()),
        "name": data.name.strip(),
        "portfolio_id": data.portfolio_id,
        "principal": data.principal,
        "interest_rate": data.interest_rate,
        "term_years": data.term_years,
        "start_date": data.start_date,
        "currency": data.currency.upper().strip(),
        "property_value": data.property_value,
        "extra_payments": 0,
        "notes": data.notes,
        "created_at": now,
        "updated_at": now,
    }
    await db.mortgages.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.put("/mortgages/{mortgage_id}", response_model=MortgageResponse)
async def update_mortgage(mortgage_id: str, data: MortgageUpdate):
    existing = await db.mortgages.find_one({"id": mortgage_id}, {"_id": 0})
    if not existing:
        raise HTTPException(status_code=404, detail="Mortgage not found")
    update_fields = {"updated_at": datetime.now(timezone.utc).isoformat()}
    for field in ["name", "principal", "interest_rate", "term_years", "start_date", "currency", "property_value", "extra_payments", "notes"]:
        val = getattr(data, field, None)
        if val is not None:
            update_fields[field] = val.strip() if isinstance(val, str) else val
    await db.mortgages.update_one({"id": mortgage_id}, {"$set": update_fields})
    updated = await db.mortgages.find_one({"id": mortgage_id}, {"_id": 0})
    return updated

@api_router.delete("/mortgages/{mortgage_id}")
async def delete_mortgage(mortgage_id: str):
    result = await db.mortgages.delete_one({"id": mortgage_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Mortgage not found")
    return {"message": "Mortgage deleted"}

# ── Universal File Import (CSV, Excel, PDF, TXT) ────────

async def parse_holdings_from_text(text: str) -> List[Dict]:
    """Parse holdings from any text format - CSV, tab-separated, or plain text."""
    holdings = []
    lines = text.strip().split('\n')
    
    # Try to detect format and parse
    for line in lines:
        line = line.strip()
        if not line:
            continue
        
        # Try different delimiters
        for delimiter in [',', '\t', '|', ';']:
            parts = [p.strip() for p in line.split(delimiter)]
            if len(parts) >= 2:
                # Try to extract symbol and shares
                symbol = None
                shares = None
                avg_price = 0
                
                for part in parts:
                    part = part.strip().upper()
                    # Check if it looks like a symbol (1-10 chars, alphanumeric with maybe - or .)
                    if re.match(r'^[A-Z]{1,10}(-[A-Z]{1,5})?(\.[A-Z]{1,3})?$', part):
                        if symbol is None:
                            symbol = part
                    else:
                        # Try to parse as number
                        clean = re.sub(r'[,$]', '', part)
                        try:
                            num = float(clean)
                            if shares is None:
                                shares = num
                            elif avg_price == 0:
                                avg_price = num
                        except:
                            pass
                
                if symbol and shares and shares > 0:
                    holdings.append({"symbol": symbol, "shares": shares, "avg_price": avg_price})
                    break
    
    return holdings

@api_router.post("/holdings/import")
async def import_file(portfolio_id: str = Form(...), file: UploadFile = File(...)):
    """Universal file import - supports CSV, Excel, PDF, and TXT formats."""
    portfolio = await db.portfolios.find_one({"id": portfolio_id}, {"_id": 0})
    if not portfolio:
        raise HTTPException(status_code=404, detail="Portfolio not found")

    content = await file.read()
    filename = (file.filename or "").lower()
    
    holdings_data = []
    message = None
    
    # Detect file type and parse
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        # Excel file
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            
            if rows:
                headers = [str(h).strip().lower() if h else "" for h in rows[0]]
                
                def find_col(names):
                    for name in names:
                        if name in headers:
                            return headers.index(name)
                    return -1

                symbol_idx = find_col(["symbol", "ticker", "stock", "name", "security"])
                shares_idx = find_col(["shares", "quantity", "qty", "units", "amount"])
                price_idx = find_col(["avg_price", "avg price", "average price", "cost", "price", "book cost per share", "avg cost"])

                if symbol_idx == -1:
                    symbol_idx = 0
                if shares_idx == -1:
                    shares_idx = 1 if len(headers) > 1 else -1

                for row in rows[1:]:
                    if not row or len(row) <= max(symbol_idx, shares_idx if shares_idx >= 0 else 0):
                        continue
                    
                    symbol = str(row[symbol_idx] or "").strip().upper()
                    if not symbol or len(symbol) > 15:
                        continue
                    
                    shares = 0
                    if shares_idx >= 0 and len(row) > shares_idx and row[shares_idx]:
                        try:
                            shares = float(str(row[shares_idx]).replace(",", "").replace("$", ""))
                        except:
                            pass
                    
                    avg_price = 0
                    if price_idx >= 0 and len(row) > price_idx and row[price_idx]:
                        try:
                            avg_price = float(str(row[price_idx]).replace(",", "").replace("$", ""))
                        except:
                            pass
                    
                    if shares > 0:
                        holdings_data.append({"symbol": symbol, "shares": shares, "avg_price": avg_price})
            
            wb.close()
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not read Excel file: {str(e)}")
    
    elif filename.endswith('.pdf'):
        # PDF file - use AI parsing
        try:
            pdf = pdfplumber.open(io.BytesIO(content))
            all_text = ""
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    all_text += page_text + "\n"
            pdf.close()
            
            if not all_text.strip():
                return {
                    "imported_count": 0,
                    "holdings": [],
                    "message": "Could not extract text from PDF. The file may be image-based or encrypted."
                }
            
            # Use AI to parse
            holdings_data = await _ai_parse_holdings(all_text)
            
            if not holdings_data:
                # Try simple text parsing as fallback
                holdings_data = await parse_holdings_from_text(all_text)
                
            if not holdings_data:
                return {
                    "imported_count": 0,
                    "holdings": [],
                    "raw_text_preview": all_text[:1000],
                    "message": "Could not detect holdings in PDF. Try CSV/Excel format or add manually."
                }
        except Exception as e:
            logger.error(f"PDF parsing error: {e}")
            raise HTTPException(status_code=400, detail=f"Could not read PDF: {str(e)}")
    
    else:
        # CSV or TXT file
        try:
            text = content.decode("utf-8")
        except:
            try:
                text = content.decode("latin-1")
            except:
                raise HTTPException(status_code=400, detail="Could not decode file. Please use UTF-8 encoding.")
        
        # Try CSV parsing first
        try:
            reader = csv.DictReader(io.StringIO(text))
            fieldnames = [f.strip().lower() for f in (reader.fieldnames or [])]
            reader.fieldnames = fieldnames
            
            for row in reader:
                symbol = (row.get("symbol") or row.get("ticker") or row.get("stock") or row.get("name") or row.get("security") or "").strip().upper()
                shares_str = (row.get("shares") or row.get("quantity") or row.get("qty") or row.get("units") or row.get("amount") or "").strip()
                price_str = (row.get("avg_price") or row.get("avg price") or row.get("average price") or row.get("cost") or row.get("price") or row.get("book cost per share") or "").strip()
                
                if not symbol or not shares_str:
                    continue
                
                shares_str = re.sub(r'[,$]', '', shares_str)
                price_str = re.sub(r'[,$]', '', price_str) if price_str else "0"
                
                try:
                    shares = float(shares_str)
                    avg_price = float(price_str) if price_str else 0
                    if shares > 0:
                        holdings_data.append({"symbol": symbol, "shares": shares, "avg_price": avg_price})
                except:
                    pass
        except:
            # Fallback to text parsing
            holdings_data = await parse_holdings_from_text(text)
    
    # Import holdings with auto-detection
    imported = []
    now = datetime.now(timezone.utc).isoformat()
    
    for h in holdings_data:
        symbol = h["symbol"]
        
        # Auto-detect ticker info
        ticker_info = await detect_ticker_info(symbol)
        
        # Check if exists
        existing = await db.holdings.find_one(
            {"symbol": ticker_info["symbol"], "portfolio_id": portfolio_id}, {"_id": 0}
        )
        
        if existing:
            await db.holdings.update_one(
                {"id": existing["id"]},
                {"$set": {
                    "shares": h["shares"],
                    "avg_price": h["avg_price"] if h["avg_price"] > 0 else existing["avg_price"],
                    "currency": ticker_info["currency"],
                    "asset_type": ticker_info["asset_type"],
                    "exchange": ticker_info["exchange"],
                    "updated_at": now
                }}
            )
            updated = await db.holdings.find_one({"id": existing["id"]}, {"_id": 0})
            updated["_action"] = "updated"
            imported.append(updated)
        else:
            doc = {
                "id": str(uuid.uuid4()),
                "symbol": ticker_info["symbol"],
                "shares": h["shares"],
                "avg_price": h["avg_price"],
                "currency": ticker_info["currency"],
                "asset_type": ticker_info["asset_type"],
                "exchange": ticker_info["exchange"],
                "notes": None,
                "portfolio_id": portfolio_id,
                "created_at": now,
                "updated_at": now,
            }
            await db.holdings.insert_one(doc)
            doc.pop("_id", None)
            doc["_action"] = "created"
            imported.append(doc)
    
    return {"imported_count": len(imported), "holdings": imported, "message": message}

# Keep old endpoints for backward compatibility
@api_router.post("/holdings/import-csv")
async def import_csv(portfolio_id: str = Form(...), file: UploadFile = File(...)):
    return await import_file(portfolio_id=portfolio_id, file=file)

@api_router.post("/holdings/import-excel")
async def import_excel(portfolio_id: str = Form(...), file: UploadFile = File(...)):
    return await import_file(portfolio_id=portfolio_id, file=file)

@api_router.post("/holdings/import-pdf")
async def import_pdf(portfolio_id: str = Form(...), file: UploadFile = File(...)):
    return await import_file(portfolio_id=portfolio_id, file=file)

# ── AI-Powered PDF Parsing ──────────────────────────────

async def _ai_parse_holdings(text: str) -> List[Dict]:
    """Use GPT to intelligently extract holdings from any document text."""
    api_key = os.environ.get('EMERGENT_LLM_KEY')
    if not api_key:
        logger.warning("EMERGENT_LLM_KEY not set - PDF AI parsing disabled")
        return []

    try:
        chat = LlmChat(
            api_key=api_key,
            session_id=f"pdf-parse-{uuid.uuid4()}",
            system_message="""You are a financial document parser. Extract ALL stock/ETF/crypto/mutual fund holdings from brokerage statements, account summaries, or portfolio reports.

Return ONLY a valid JSON array. Each item must have:
- "symbol": ticker symbol (e.g. "AAPL", "QQQ", "SHOP.TO", "BTC-USD", "RELIANCE.NS") - uppercase, include exchange suffix if present
- "shares": number of shares/units (float)
- "avg_price": average cost per share (float, use 0 if not found)

IMPORTANT RULES:
1. Extract ALL holdings you can find - stocks, ETFs, mutual funds, crypto
2. For Canadian stocks, keep the .TO or .V suffix if present
3. For Indian stocks, keep the .NS or .BO suffix if present
4. For crypto, use format like BTC-USD, ETH-USD
5. If you see "quantity", "units", "shares" - that's the shares count
6. If you see "book cost", "avg cost", "average price" - that's the avg_price
7. Ignore cash balances, totals, and non-holding entries
8. Return EMPTY array [] if no holdings found
9. Return ONLY the JSON array - no explanation"""
        ).with_model("openai", "gpt-4o")

        # Truncate text to avoid token limits
        truncated_text = text[:12000] if len(text) > 12000 else text
        
        user_msg = UserMessage(
            text=f"Extract ALL stock/ETF/crypto holdings from this document. Return JSON array:\n\n{truncated_text}"
        )
        
        response = await chat.send_message(user_msg)
        response_text = response.strip()
        
        # Handle markdown code blocks
        if "```" in response_text:
            match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', response_text)
            if match:
                response_text = match.group(1).strip()
        
        # Try to find JSON array in response
        if not response_text.startswith('['):
            match = re.search(r'\[[\s\S]*\]', response_text)
            if match:
                response_text = match.group(0)
        
        parsed = json_module.loads(response_text)
        if not isinstance(parsed, list):
            return []

        # Validate entries
        results = []
        for item in parsed:
            sym = str(item.get("symbol", "")).strip().upper()
            shares = float(item.get("shares", 0))
            avg_price = float(item.get("avg_price", 0))
            
            # Basic validation
            if sym and shares > 0 and len(sym) <= 15:
                results.append({
                    "symbol": sym,
                    "shares": shares,
                    "avg_price": avg_price
                })
        
        logger.info(f"AI parsed {len(results)} holdings from PDF")
        return results
        
    except Exception as e:
        logger.error(f"AI parsing error: {e}")
        return []

# ── Stock Quotes (FREE via yfinance) ────────────────────

@api_router.get("/fx-rates")
async def get_fx_rates():
    """Fetch live FX rates using yfinance (FREE)."""
    rates = {"USD": 1.0}
    pairs = {"CAD": "USDCAD=X", "INR": "USDINR=X", "GBP": "USDGBP=X", "EUR": "USDEUR=X"}
    for currency, symbol in pairs.items():
        try:
            ticker = yf.Ticker(symbol)
            info = ticker.fast_info
            rate = float(info.last_price) if hasattr(info, 'last_price') and info.last_price else 0
            rates[currency] = round(rate, 4)
        except Exception as e:
            logger.warning(f"Error fetching FX rate {symbol}: {e}")
            # Fallback rates
            fallbacks = {"CAD": 1.36, "INR": 84.5, "GBP": 0.79, "EUR": 0.92}
            rates[currency] = fallbacks.get(currency, 1)
    return rates

@api_router.get("/stocks/quotes")
async def get_stock_quotes(symbols: str, currencies: str = ""):
    """Fetch live quotes using yfinance (FREE)."""
    symbol_list = [s.strip().upper() for s in symbols.split(",") if s.strip()]
    if not symbol_list:
        return {}

    quotes = {}
    
    for sym in symbol_list:
        try:
            # Use symbol as-is first (it might already have exchange suffix)
            ticker = yf.Ticker(sym)
            info = ticker.fast_info
            price = round(float(info.last_price), 2) if hasattr(info, 'last_price') and info.last_price else 0
            
            # Get currency from ticker
            price_currency = "USD"
            try:
                ticker_info = ticker.info
                price_currency = ticker_info.get("currency", "USD").upper()
            except:
                pass
            
            quotes[sym] = {
                "price": price,
                "previous_close": round(float(info.previous_close), 2) if hasattr(info, 'previous_close') and info.previous_close else 0,
                "day_high": round(float(info.day_high), 2) if hasattr(info, 'day_high') and info.day_high else 0,
                "day_low": round(float(info.day_low), 2) if hasattr(info, 'day_low') and info.day_low else 0,
                "market_cap": float(info.market_cap) if hasattr(info, 'market_cap') and info.market_cap else 0,
                "quote_currency": price_currency,
            }
        except Exception as e:
            logger.warning(f"Error fetching quote for {sym}: {e}")
            quotes[sym] = {"price": 0, "previous_close": 0, "day_high": 0, "day_low": 0, "market_cap": 0, "quote_currency": "USD"}
    
    return quotes

@api_router.get("/stocks/history/{symbol}")
async def get_stock_history(symbol: str, period: str = "1mo"):
    """Fetch historical data using yfinance (FREE)."""
    valid_periods = ["1d", "5d", "1mo", "3mo", "6mo", "1y", "2y", "5y"]
    if period not in valid_periods:
        raise HTTPException(status_code=400, detail=f"Invalid period. Use one of: {valid_periods}")
    try:
        ticker = yf.Ticker(symbol.upper())
        hist = ticker.history(period=period)
        if hist.empty:
            return {"symbol": symbol.upper(), "data": []}
        data_points = []
        for date, row in hist.iterrows():
            data_points.append({
                "date": date.strftime("%Y-%m-%d"),
                "close": round(float(row["Close"]), 2),
                "high": round(float(row["High"]), 2),
                "low": round(float(row["Low"]), 2),
                "open": round(float(row["Open"]), 2),
                "volume": int(row["Volume"]),
            })
        return {"symbol": symbol.upper(), "period": period, "data": data_points}
    except Exception as e:
        logger.error(f"Error fetching history for {symbol}: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ── Asset Types & Summary ───────────────────────────────

@api_router.get("/asset-types")
async def get_asset_types():
    return {"asset_types": ASSET_TYPES}

@api_router.get("/portfolio/summary")
async def get_portfolio_summary(portfolio_id: Optional[str] = None):
    query = {}
    if portfolio_id:
        query["portfolio_id"] = portfolio_id
    
    holdings = await db.holdings.find(query, {"_id": 0}).to_list(500)
    mortgages = await db.mortgages.find(query, {"_id": 0}).to_list(100)
    
    summary = {
        "total_holdings": len(holdings),
        "total_mortgages": len(mortgages),
        "by_asset_type": {},
        "by_currency": {},
        "by_exchange": {},
    }
    
    for h in holdings:
        asset_type = h.get("asset_type", "Stock")
        currency = h.get("currency", "USD")
        exchange = h.get("exchange") or "Unknown"
        
        if asset_type not in summary["by_asset_type"]:
            summary["by_asset_type"][asset_type] = {"count": 0, "total_value": 0}
        summary["by_asset_type"][asset_type]["count"] += 1
        summary["by_asset_type"][asset_type]["total_value"] += h["shares"] * h["avg_price"]
        
        if currency not in summary["by_currency"]:
            summary["by_currency"][currency] = {"count": 0}
        summary["by_currency"][currency]["count"] += 1
        
        if exchange not in summary["by_exchange"]:
            summary["by_exchange"][exchange] = {"count": 0}
        summary["by_exchange"][exchange]["count"] += 1
    
    return summary

# ── Health ──────────────────────────────────────────────

@api_router.get("/")
async def root():
    return {"message": "Holdings Hub API - 100% FREE", "features": ["auto-detect-exchange", "auto-detect-currency", "multi-format-import"]}

@api_router.get("/health")
async def health():
    return {"status": "healthy", "free": True}

# ── Include Router & Middleware ─────────────────────────

app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
